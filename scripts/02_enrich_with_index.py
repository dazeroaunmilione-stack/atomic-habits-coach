import json
import openpyxl

# Carica moduli estratti
with open('../output/modules_raw.json', 'r', encoding='utf-8') as f:
    modules = json.load(f)

# Carica Indice Strutturale - sheet post_gate_retrieval
wb = openpyxl.load_workbook('../files/Indice_Strutturale_Moduli_v2.4.xlsx')
ws = wb['indice_post_gate_retrieval']

# Leggi headers
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]

# Costruisci dizionario moduli per module_id e titolo
index_data = {}
for r in range(2, ws.max_row+1):
    row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column+1)}
    if row.get('module_id'):
        index_data[str(row['module_id'])] = row
        if row.get('module_title'):
            index_data[row['module_title'].lower().strip()] = row

# Arricchisci moduli con metadati
enriched = []
for i, mod in enumerate(modules):
    meta = index_data.get(str(i+1)) or index_data.get(mod['title'].lower().strip(), {})
    enriched_mod = {
        'module_id': meta.get('module_id', i+1),
        'module_title': meta.get('module_title', mod['title']),
        'canonical_label': meta.get('canonical_label', ''),
        'primary_domain': meta.get('primary_domain', ''),
        'subdomain': meta.get('subdomain', ''),
        'topic_type': meta.get('topic_type', ''),
        'functional_role': meta.get('functional_role', ''),
        'behavior_change_stage': meta.get('behavior_change_stage', ''),
        'coaching_relevance': meta.get('coaching_relevance', ''),
        'retrieval_keywords': meta.get('retrieval_keywords', ''),
        'post_gate_candidate_activation_signals': meta.get('post_gate_candidate_activation_signals', ''),
        'text': mod['text'],
        'page': mod['page']
    }
    enriched.append(enriched_mod)
    print(f"Modulo {i+1}: {enriched_mod['module_title'][:50]} — domain: {enriched_mod['primary_domain']}")

with open('../output/modules_enriched.json', 'w', encoding='utf-8') as f:
    json.dump(enriched, f, ensure_ascii=False, indent=2)

print(f"\n✓ Salvati {len(enriched)} moduli arricchiti in output/modules_enriched.json")