import json
import openpyxl

wb = openpyxl.load_workbook('../files/Behavior_Diagnosis_Matrix_v2.4.xlsx')

# PRE-GATE: solo campi neutri
ws_pre = wb['bdm_core_pre_gate']
headers_pre = [ws_pre.cell(1,c).value for c in range(1, ws_pre.max_column+1)]
bdm_pre = []
for r in range(2, ws_pre.max_row+1):
    row = {headers_pre[c-1]: ws_pre.cell(r,c).value for c in range(1, ws_pre.max_column+1)}
    if row.get('diagnosis_id'):
        bdm_pre.append(row)

# POST-GATE: tutti i campi
ws_post = wb['bdm_post_gate_guidance']
headers_post = [ws_post.cell(1,c).value for c in range(1, ws_post.max_column+1)]
bdm_post = []
for r in range(2, ws_post.max_row+1):
    row = {headers_post[c-1]: ws_post.cell(r,c).value for c in range(1, ws_post.max_column+1)}
    if row.get('diagnosis_id'):
        bdm_post.append(row)

with open('../output/bdm_pre_gate.json', 'w', encoding='utf-8') as f:
    json.dump(bdm_pre, f, ensure_ascii=False, indent=2)

with open('../output/bdm_post_gate.json', 'w', encoding='utf-8') as f:
    json.dump(bdm_post, f, ensure_ascii=False, indent=2)

print(f"✓ BDM pre-gate: {len(bdm_pre)} pattern")
print(f"✓ BDM post-gate: {len(bdm_post)} pattern")